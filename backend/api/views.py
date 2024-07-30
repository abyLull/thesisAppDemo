from itertools import count

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from .forms import RegistrationForm,DataUploadForm,StudentsForm, DestinationsForm
from tablib import Dataset
from .resources import StudentResource, DestinationResource
from .models import Student,Destination
from .decorators import allowed_users
import csv,io
import openpyxl
import logging

logger = logging.getLogger(__name__)

# Create your views here.


@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff', 'guest'])
def UploadStudentDataViewXLSX(request):
    if request.method == 'POST':
        student_resource = StudentResource()
        dataset = Dataset()
        new_students = request.FILES['file']
        year = request.POST.get('year')
        comment = request.POST.get('comment')

        if not new_students:
            messages.info(request, 'No file uploaded.')
            return redirect('students_xlsx')

        if not new_students.name.endswith('xlsx'):
            messages.info(request, 'Please upload an XLSX file')
            return redirect('students_xlsx')
        else:
            messages.info(request, 'File successfully uploaded')

        # Clear existing students and their destinations for the given year
        try:
            destinations_to_clear = Destination.objects.filter(year=year)
            for destination in destinations_to_clear:
                destination.student.clear()  # Clear the many-to-many relationships

            data_set = dataset.load(new_students.read(), format='xlsx')
            for column in data_set:
                created = Student(
                    id=column[0],
                    name=column[2],
                    surname=column[1],
                    year_of_study=column[8],
                    contract_codes=column[11],
                    year=year,
                    faculty=request.user.profile.faculty,
                    comment=comment,
                    study_level=column[7],
                    status='Pending',
                    gpa=0,
                    modified='negative',
                )
                created.save()
            messages.success(request, 'File uploaded and students added successfully.')
            return redirect('students_xlsx')
        except Exception as e:
            logger.error(f'Error processing file: {str(e)}', exc_info=True)
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('students_xlsx')

    # Pagination logic
    student_list = Student.objects.all().order_by('id')
    paginator = Paginator(student_list, 10)  # Show 10 students per page
    page_number = request.GET.get('page')
    student_page_obj = paginator.get_page(page_number)

    return render(request, 'studentsXLSX.html', {'student_page_obj': student_page_obj})


@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff', 'guest'])
def UploadDestinationDataViewXLSX(request):
    if request.method == 'POST':
        destination_resource = DestinationResource()
        dataset = Dataset()
        new_destinations = request.FILES['file']
        year = request.POST.get('year')
        comment = request.POST.get('comment')

        if not new_destinations.name.endswith('xlsx'):
            messages.info(request, 'Please upload an XLSX file')
            return render(request, 'destinationsXLSX.html')
        else:
            messages.info(request, 'File successfully uploaded')

        destinations_to_clear = Destination.objects.filter(year=year)
        for destination in destinations_to_clear:
            destination.student.clear()
        destinations_to_clear.delete()

        data_set = dataset.load(new_destinations.read(), format='xlsx')
        for column in data_set:
            created = Destination(
                id=column[0],
                country=column[5],
                university=column[3],
                available_places=column[16],
                contract_code=column[1],
                year=year,
                faculty=request.user.profile.faculty,
                study_level_available=column[15],
                comment=comment,
            )
            created.save()

    destinations_list = Destination.objects.all().order_by('id')

    # Pagination
    paginator = Paginator(destinations_list, 10)  # Show 10 destinations per page
    page_number = request.GET.get('page')
    destinations_page_obj = paginator.get_page(page_number)

    return render(request, 'destinationsXLSX.html', {'destinations_page_obj': destinations_page_obj})
@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff'])
def edit_destination(request):
    if request.method == 'POST':
        destination_id = request.POST.get('id')
        destination = get_object_or_404(Destination, id=destination_id)
        form = DestinationsForm(request.POST, instance=destination)
        if form.is_valid():
            form.save()
            return redirect('destinations_xlsx')

@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff'])
def delete_destination(request):
    if request.method == 'POST':
        destination_ids = request.POST.getlist('destination_ids')
        if destination_ids:
            destinations = Destination.objects.filter(pk__in=destination_ids)
            destinations.delete()
            messages.success(request, 'Selected destinations deleted successfully.')
        else:
            messages.error(request, 'No destinations selected for deletion.')
        return redirect('destinations_xlsx')

@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff'])
def edit_student(request):
    if request.method == 'POST':
        student_id = request.POST.get('id')
        student = get_object_or_404(Student, pk=student_id)
        form = StudentsForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect('students_xlsx')

@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff'])
def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    data = {
        'id': student.id,
        'name': student.name,
        'surname': student.surname,
        'year_of_study': student.year_of_study,
        'study_level': student.study_level,
        'contract_codes': student.contract_codes,
        'status': student.status,
        'year': student.year,
        'comment': student.comment,
        'gpa': student.gpa,
        'modified': student.modified,
    }
    return JsonResponse(data)

@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff'])
def destination_detail(request, destination_id):
    destination = get_object_or_404(Destination, id=destination_id)
    data = {
        'id': destination.id,
        'country': destination.country,
        'university': destination.university,
        'available_places': destination.available_places,
        'study_level_available': destination.study_level_available,
        'contract_code': destination.contract_code,
        'year': destination.year,
        'comment': destination.comment,
    }
    return JsonResponse(data)
@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff'])
def delete_student(request):
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        if student_ids:
            students = Student.objects.filter(pk__in=student_ids)
            students.delete()
            messages.success(request, 'Selected students deleted successfully.')
        else:
            messages.error(request, 'No students selected for deletion.')
        return redirect('students_xlsx')


@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff'])
def add_student(request):
    if request.method == 'POST':
        indeks = request.POST['id']
        name = request.POST['name']
        surname = request.POST['surname']
        year_of_study = request.POST['year_of_study']
        contract_codes = request.POST['contract_codes']
        year = request.POST['year']
        gpa = request.POST['gpa']
        modified = request.POST['modified']
        study_level = request.POST['study_level']
        comment = request.POST['comment']
        faculty = request.user.profile.faculty

        student = Student(
            id=indeks,
            name=name,
            surname=surname,
            year_of_study=year_of_study,
            study_level=study_level,
            contract_codes=contract_codes,
            year=year,
            comment=comment,
            faculty=faculty,
            gpa=gpa,
            modified=modified
        )
        student.save()
        return redirect('students_xlsx')

@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff'])
def add_destination(request):
    if request.method == 'POST':
        country = request.POST['country']
        university = request.POST['university']
        available_places = request.POST['available_places']
        contract_code = request.POST['contract_code']
        year = request.POST['year']
        faculty = request.user.profile.faculty
        study_level_available= request.POST['study_level_available']

        destination = Destination(
            country=country,
            university=university,
            available_places=available_places,
            study_level_available=study_level_available,
            contract_code=contract_code,
            year=year,
            faculty=faculty
        )
        destination.save()
        return redirect('destinations_xlsx')

def Base(request):
    return render(request, 'main/base.html')
def Home(request):
    return render(request, 'main/home.html')

def sign_up(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.profile.faculty = form.cleaned_data['faculty']
            user.profile.save()
            user = authenticate(username=form.cleaned_data['username'], password=form.cleaned_data['password1'])
            group = Group.objects.get(name = 'guest')
            user.groups.add(group)
            login(request, user)
            return redirect('/home')
    else:
        form = RegistrationForm()

    return render(request, 'registration/sign_up.html', {"form": form})
@login_required(login_url="/login")
def logout_view(request):
    if request.method == 'POST':
        logout(request)
        return redirect("/login")
    return render(request,"registration/logout_user.html")

@login_required(login_url="/login")
def SortedListView(request):
    if request.method == 'POST':
        destination_id = request.POST.get('destination_id')
        try:
            destination = Destination.objects.get(id=destination_id)
        except Destination.DoesNotExist:
            messages.error(request, "Destination not found.")
            return redirect('sorted_list')

        if 'delete_students' in request.POST:
            student_ids = request.POST.getlist('delete_student_ids')
            for student_id in student_ids:
                try:
                    student = Student.objects.get(id=student_id)
                    destination.student.remove(student)
                    destination.save()
                    student.status = 'Pending'
                    student.save()
                except Student.DoesNotExist:
                    messages.error(request, f"Student with id {student_id} not found.")
            return redirect('sorted_list')

        if 'add_students' in request.POST:
            student_ids = request.POST.getlist('add_student_ids')
            available_places = destination.available_places

            if len(student_ids) > available_places:
                messages.warning(request, f'Too many students selected for {destination.university} ({destination.contract_code}).')
            else:
                for student_id in student_ids:
                    try:
                        student = Student.objects.get(id=student_id)
                        if can_assign_student_to_destination(student, destination) and destination.student.count() < destination.available_places and destination.year == student.year:
                            destination.student.add(student)
                            student.status = 'Approved'
                            student.save()
                        else:
                            messages.warning(request, f'Cannot assign student {student.name} to {destination.university}.')
                    except Student.DoesNotExist:
                        messages.error(request, f"Student with id {student_id} not found.")
                destination.save()
            return redirect('sorted_list')

        if 'auto-assign' in request.POST:
            year = request.POST.get('year')
            if year:
                students = Student.objects.filter(year=year, status='Pending')
                destinations = Destination.objects.filter(year=year)
            else:
                students = Student.objects.filter(status='Pending')
                destinations = Destination.objects.all()

            students = students.order_by('-gpa', '-modified')
            for student in students:
                for destination in destinations:
                    if can_assign_student_to_destination(student, destination) and destination.student.count() < destination.available_places and destination.year == student.year:
                        destination.student.add(student)
                        student.status = 'Approved'
                        student.save()
                        destination.save()
                        break
            return redirect('sorted_list')

    year = request.GET.get('year')
    if year:
        destinations = Destination.objects.filter(year=year).order_by('id')
        students = Student.objects.filter(status='Pending', year=year).order_by('id')
    else:
        destinations = Destination.objects.all().order_by('id')
        students = Student.objects.filter(status='Pending').order_by('id')

    paginator = Paginator(destinations, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    destination_student_mapping = {}
    for destination in destinations:
        matching_students = [
            student for student in students
            if destination.contract_code in student.contract_codes.split(';') and
            student.year == destination.year and
            can_assign_student_to_destination(student, destination) and student.status == 'Pending'
        ]
        destination_student_mapping[destination.id] = matching_students

    logger.debug("destination_student_mapping: %s", destination_student_mapping)

    return render(request, 'mobility/sorted.html', {
        'page_obj': page_obj,
        'destination_student_mapping': destination_student_mapping,
        'students': students,
    })

def can_assign_student_to_destination(student, destination):
    study_level_match = {
        'pierwszego stopnia': ['pierwszy, drugi i trzeci', 'pierwszego stopnia', 'pierwszy i drugi'],
        'drugiego stopnia': ['pierwszy, drugi i trzeci', 'drugiego stopnia', 'pierwszy i drugi'],
        'trzeciego stopnia': ['pierwszy, drugi i trzeci', 'trzeciego stopnia'],
    }
    return destination.contract_code in student.contract_codes.split(';') and \
           destination.study_level_available in study_level_match.get(student.study_level, [])



@login_required(login_url="/login")
#@allowed_users(allowed_roles=['staff', 'guest'])
def ApprovedListView(request):
    year = request.GET.get('year')
    if year:
        destinations = Destination.objects.filter(year=year).order_by('id')
        approved_students = Student.objects.filter(status='Approved', year=year).order_by('id')
    else:
        destinations = Destination.objects.all().order_by('id')
        approved_students = Student.objects.filter(status='Approved').order_by('id')

    destination_paginator = Paginator(destinations, 10)
    destination_page_number = request.GET.get('destination_page', 1)
    destination_page_obj = destination_paginator.get_page(destination_page_number)

    student_paginator = Paginator(approved_students, 10)
    student_page_number = request.GET.get('student_page', 1)
    student_page_obj = student_paginator.get_page(student_page_number)

    destination_student_mapping = {}
    for destination in destinations:
        matching_students = [
            student for student in approved_students
            if destination.contract_code in student.contract_codes.split(';')
        ]
        destination_student_mapping[destination.id] = matching_students

    student_destination_mapping = {}
    for student in approved_students:
        for destination in destinations:
            if student in destination.student.all():
                student_destination_mapping[student.id] = destination
                break

    if request.GET.get('download_approved_csv'):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="approved_students_{year if year else "all"}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Name', 'Surname', 'Year of Study', 'Country', 'University'])
        for student in approved_students:
            destination = student_destination_mapping.get(student.id, None)
            writer.writerow([
                student.name,
                student.surname,
                student.year_of_study,
                destination.country if destination else 'No destination assigned',
                destination.university if destination else 'No destination assigned'
            ])
        return response

    if request.GET.get('download_approved_xlsx'):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="approved_students_{year if year else "all"}.xlsx"'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Approved Students'

        sheet.append(['Name', 'Surname', 'Year of Study', 'Country', 'University'])
        for student in approved_students:
            destination = student_destination_mapping.get(student.id, None)
            sheet.append([
                student.name,
                student.surname,
                student.year_of_study,
                destination.country if destination else 'No destination assigned',
                destination.university if destination else 'No destination assigned'
            ])

        workbook.save(response)
        return response

    if request.GET.get('download_destinations_csv'):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="destinations_{year if year else "all"}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Id', 'Country', 'University', 'Available Places', 'Contract Code'])
        for destination in destinations:
            writer.writerow([
                destination.id,
                destination.country,
                destination.university,
                destination.available_places,
                destination.contract_code,
            ])
        return response

    if request.GET.get('download_destinations_xlsx'):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="destinations_{year if year else "all"}.xlsx"'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Destinations'

        sheet.append(['Id', 'Country', 'University', 'Available Places', 'Contract Code', 'Assigned Student'])
        for destination in destinations:
            matching_students = [
                student for student in approved_students
                if destination.contract_code in student.contract_codes.split(';')
            ]
            student_names = ', '.join([f"{student.name} {student.surname}" for student in matching_students])
            sheet.append([
                destination.id,
                destination.country,
                destination.university,
                destination.available_places,
                destination.contract_code,
                student_names,
            ])

        workbook.save(response)
        return response

    context = {
        'destination_page_obj': destination_page_obj,
        'student_page_obj': student_page_obj,
        'destination_student_mapping': destination_student_mapping,
        'approved_students': approved_students,
        'student_destination_mapping': student_destination_mapping,
    }

    return render(request, 'mobility/approved.html', context)
#GPA not impoeted from file. same student cant be assigned to multiple destinations only one)
@login_required(login_url="/login")
def download_approved_xlsx(request):
    year = request.GET.get('year')
    if year:
        destinations = Destination.objects.filter(faculty=request.user.profile.faculty, year=year)
    else:
        destinations = Destination.objects.filter(faculty=request.user.profile.faculty)

    approved_students = Student.objects.filter(status='Approved')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=approved_list.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved List"

    ws.append(['Country', 'University', 'Contract Code', 'Available Places', 'Student(s)'])

    for destination in destinations:
        matching_students = [
            student for student in approved_students
            if destination.contract_code in student.contract_codes.split(';')
        ]
        student_names = ', '.join([f"{student.name} {student.surname}" for student in matching_students])
        ws.append([destination.country, destination.university, destination.contract_code, destination.available_places,
                   student_names])

    wb.save(response)
    return response
@login_required(login_url="/login")
def view_destination(request, id):
    destination = get_object_or_404(Destination, id=id)
    return render(request, 'mobility/view_destination.html', {'destination': destination})



